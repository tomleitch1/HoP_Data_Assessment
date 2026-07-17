"""
Parliament Finance Systems Programme
Full DQ Export Generator
========================
Generates a DQ tracker AND individual evidence files (one per failing check)
for a specific tab and house.

Usage:
    python scripts/generate_full_export.py suppliers HOC
    python scripts/generate_full_export.py suppliers HOL
    python scripts/generate_full_export.py gl HOC
    python scripts/generate_full_export.py assets HOL
    python scripts/generate_full_export.py po HOC

Tracker saved to:  trackers/<tab>_tracker_<HOUSE>.xlsx  (same as generate_tracker.py)
Evidence saved to: trackers/evidence/<HOUSE>_<Dimension>_<CHECK_ID>.xlsx
                   (only checks with at least one failing record)

Valid tabs:   suppliers, customers, gl, assets, po
Valid houses: HOC, HOL (po is HoC-only — running po HOL will simply find no data)
"""

import sys
import os
import argparse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dashboard.data_engine import load_data, run_dq_analysis, get_failing_records
from dashboard.core.config import SCOPE_CONFIG

TAB_ALIASES = {
    'suppliers': 'suppliers',
    'ap':        'suppliers',
    'customers': 'customers',
    'ar':        'customers',
    'gl':        'gl',
    'assets':    'assets',
    'po':        'po',
}

# Maps engine tab key → SCOPE_CONFIG key
SCOPE_KEY = {
    'suppliers': 'ap',
    'customers': 'ar',
    'gl':        'gl',
    'assets':    'assets',
    'po':        'po',
}

HEADER_FILL  = PatternFill('solid', fgColor='1E1528')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=10)
THIN_BORDER  = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
)
SEV_FILL = {
    'Critical': 'FFD7D7',
    'High':     'FFE8CC',
    'Medium':   'FFF3CC',
    'Low':      'E8F5E9',
}
RAG_FILL = {
    'Red':   'FFCCCC',
    'Amber': 'FFE5B4',
    'Green': 'CCFFCC',
}


# ---------------------------------------------------------------------------
# Tracker (mirrors generate_tracker.py exactly)
# ---------------------------------------------------------------------------

def _write_tracker(out_df: pd.DataFrame, out_path: str) -> None:
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='DQ Tracker')
        ws = writer.sheets['DQ Tracker']

        col_widths = {
            'A': 30, 'B': 55, 'C': 20, 'D': 12,
            'E': 18, 'F': 16, 'G': 12, 'H': 8,
            'I': 40, 'J': 35,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30

        sev_col  = out_df.columns.get_loc('Severity') + 1
        rag_col  = out_df.columns.get_loc('RAG') + 1
        fail_col = out_df.columns.get_loc('Failing Records') + 1

        for i, (_, row) in enumerate(out_df.iterrows(), start=2):
            ws.row_dimensions[i].height = 18
            for cell in ws[i]:
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=False)
                cell.font      = Font(size=10)

            sev_fill = SEV_FILL.get(row['Severity'])
            if sev_fill:
                sev_cell = ws.cell(row=i, column=sev_col)
                sev_cell.fill = PatternFill('solid', fgColor=sev_fill)
                sev_cell.font = Font(bold=True, size=10)

            rag_fill = RAG_FILL.get(row['RAG'])
            if rag_fill:
                rag_cell = ws.cell(row=i, column=rag_col)
                rag_cell.fill = PatternFill('solid', fgColor=rag_fill)
                rag_cell.font = Font(bold=True, size=10)

            ws.cell(row=i, column=fail_col).font = Font(bold=True, size=10)

        ws.freeze_panes = 'A2'


def make_tracker(tab_key: str, house: str, dq_results: pd.DataFrame, out_dir: str) -> None:
    scope_ids = SCOPE_CONFIG[SCOPE_KEY[tab_key]]['scope_ids']

    df = dq_results[
        (dq_results['scope_id'].isin(scope_ids)) &
        (dq_results['house'] == house) &
        (dq_results['failing'] > 0)
    ].copy()

    if df.empty:
        print(f"  No failing checks for tracker — skipping.")
        return

    df = df.sort_values(['scope_id', 'severity', 'check_id'])

    rows = []
    for _, row in df.iterrows():
        rows.append({
            'Test Reference':                  row['check_id'],
            'Description':                     row['description'],
            'Dimension':                       row['dimension'],
            'Severity':                        row['severity'],
            'Failing Records':                 int(row['failing']),
            'Total Assessed':                  int(row['total']),
            'Error Rate':                      f"{row['error_rate']:.1f}%",
            'RAG':                             row['rag'],
            'Comments':                        '',
            'Source System Cleansing Complete': '',
        })

    out_df  = pd.DataFrame(rows)
    out_path = os.path.join(out_dir, f"{tab_key}_tracker_{house}.xlsx")
    _write_tracker(out_df, out_path)
    print(f"  Tracker saved: {out_path}  ({len(out_df)} checks with failures)")


# ---------------------------------------------------------------------------
# Evidence files
# ---------------------------------------------------------------------------

def _style_evidence_sheet(ws, df: pd.DataFrame) -> None:
    """Apply consistent styling to an evidence worksheet."""
    # Auto-width (capped at 60)
    for col_idx, col_name in enumerate(df.columns, start=1):
        sample = df[col_name].head(200) if not df.empty else df[col_name]
        max_len = max(len(str(col_name)), sample.astype(str).str.len().max() if not sample.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Data rows
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 16
        for cell in ws[i]:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.font      = Font(size=9)

    ws.freeze_panes = 'A2'


def _safe_filename(house: str, dimension: str, check_id: str) -> str:
    return f"{house}_{dimension}_{check_id}.xlsx"


def export_evidence(tab_key: str, house: str, dq_results: pd.DataFrame,
                    frames: dict, out_dir: str) -> None:
    scope_ids = SCOPE_CONFIG[SCOPE_KEY[tab_key]]['scope_ids']

    failing_checks = dq_results[
        (dq_results['scope_id'].isin(scope_ids)) &
        (dq_results['house'] == house) &
        (dq_results['failing'] > 0)
    ].copy()

    if failing_checks.empty:
        print(f"  No failing checks — no evidence files to write.")
        return

    written = 0
    for _, row in failing_checks.iterrows():
        check_id  = row['check_id']
        dimension = row['dimension']

        try:
            records = get_failing_records(check_id, house, frames, for_export=True)
        except Exception as exc:
            print(f"  [WARN] Could not retrieve records for {check_id}: {exc}")
            continue

        if records is None or records.empty:
            continue

        fname    = _safe_filename(house, dimension, check_id)
        out_path = os.path.join(out_dir, fname)

        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            # Meta sheet
            meta_rows = [
                ['Check ID',       check_id],
                ['Description',    row['description']],
                ['Dimension',      dimension],
                ['Severity',       row['severity']],
                ['RAG',            row['rag']],
                ['Failing Records', int(row['failing'])],
                ['Total Assessed', int(row['total'])],
                ['Error Rate',     f"{row['error_rate']:.1f}%"],
                ['House',          house],
            ]
            meta_df = pd.DataFrame(meta_rows, columns=['Field', 'Value'])
            meta_df.to_excel(writer, index=False, sheet_name='Check Info')
            ws_meta = writer.sheets['Check Info']
            ws_meta.column_dimensions['A'].width = 22
            ws_meta.column_dimensions['B'].width = 70
            for cell in ws_meta[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            for i in range(2, ws_meta.max_row + 1):
                for cell in ws_meta[i]:
                    cell.border = THIN_BORDER
                    cell.font   = Font(size=10)
                ws_meta.cell(row=i, column=1).font = Font(bold=True, size=10)

            # Records sheet
            records.to_excel(writer, index=False, sheet_name='Failing Records')
            _style_evidence_sheet(writer.sheets['Failing Records'], records)

        written += 1

    print(f"  Evidence files saved: {written} files in {out_dir}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run(tab: str, house: str) -> None:
    tab_key = TAB_ALIASES.get(tab.lower())
    if tab_key is None:
        print(f"Unknown tab '{tab}'. Valid: {', '.join(TAB_ALIASES)}")
        sys.exit(1)

    house = house.upper()
    if house not in ('HOC', 'HOL'):
        print("House must be HOC or HOL.")
        sys.exit(1)

    print(f"Loading data...")
    frames = load_data(tab=tab_key)
    print(f"Running DQ analysis...")
    dq_results = run_dq_analysis(frames, tab=tab_key)

    tracker_dir  = 'trackers'
    evidence_dir = os.path.join('trackers', 'evidence')
    os.makedirs(tracker_dir,  exist_ok=True)
    os.makedirs(evidence_dir, exist_ok=True)

    print(f"\n[1/2] Generating tracker...")
    make_tracker(tab_key, house, dq_results, tracker_dir)

    print(f"\n[2/2] Generating evidence files...")
    export_evidence(tab_key, house, dq_results, frames, evidence_dir)

    print(f"\nDone.")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate DQ tracker + evidence exports')
    parser.add_argument('tab',   help='Tab name: suppliers, customers, gl, assets, po')
    parser.add_argument('house', help='House: HOC or HOL')
    args = parser.parse_args()
    run(args.tab, args.house)
