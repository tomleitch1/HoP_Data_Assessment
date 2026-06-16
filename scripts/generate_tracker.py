"""
Parliament Finance Systems Programme
DQ Tracker Generator
====================
Generates an Excel tracker for a specific tab and house, listing only checks
with at least one failing record. The tracker is pre-populated with check
metadata and has empty columns for manual completion.

Usage:
    python scripts/generate_tracker.py suppliers HOC
    python scripts/generate_tracker.py suppliers HOL
    python scripts/generate_tracker.py gl HOC
    python scripts/generate_tracker.py assets HOL

Valid tabs: suppliers, customers, gl, assets
Valid houses: HOC, HOL

Output: trackers/<tab>_tracker_<HOUSE>.xlsx
"""

import sys
import os
import argparse

# Allow imports from project root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dashboard.data_engine import load_data, run_dq_analysis
from dashboard.core.config import SCOPE_CONFIG

TAB_ALIASES = {
    'suppliers': 'suppliers',
    'ap':        'suppliers',
    'customers': 'customers',
    'ar':        'customers',
    'gl':        'gl',
    'assets':    'assets',
}

SCOPE_KEY = {
    'suppliers': 'ap',
    'customers': 'ar',
    'gl':        'gl',
    'assets':    'assets',
}

SEV_FILL = {
    'Critical': 'FFD7D7',
    'High':     'FFE8CC',
    'Medium':   'FFF3CC',
    'Low':      'E8F5E9',
}
RAG_FILL = {
    'Red':   'FFCCCC',
    'Amber': 'FFE5B4',
    'Green': 'CCFFCC',
}

HEADER_FILL   = PatternFill('solid', fgColor='1E1528')
HEADER_FONT   = Font(color='FFFFFF', bold=True, size=10)
SUBHEAD_FILL  = PatternFill('solid', fgColor='EDE9F8')
SUBHEAD_FONT  = Font(bold=True, size=10)
THIN_BORDER   = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
)


def make_tracker(tab: str, house: str) -> None:
    tab_key = TAB_ALIASES.get(tab.lower())
    if tab_key is None:
        print(f"Unknown tab '{tab}'. Valid: {', '.join(TAB_ALIASES)}")
        sys.exit(1)

    house = house.upper()
    if house not in ('HOC', 'HOL'):
        print("House must be HOC or HOL.")
        sys.exit(1)

    print(f"Loading data...")
    frames = load_data(tab=tab_key)
    print(f"Running DQ analysis...")
    dq_results = run_dq_analysis(frames, tab=tab_key)

    scope_ids = SCOPE_CONFIG[SCOPE_KEY[tab_key]]['scope_ids']

    df = dq_results[
        (dq_results['scope_id'].isin(scope_ids)) &
        (dq_results['house'] == house) &
        (dq_results['failing'] > 0)
    ].copy()

    if df.empty:
        print(f"No failing checks found for {tab} / {house}.")
        return

    df = df.sort_values(['scope_id', 'severity', 'check_id'])

    rows = []
    for _, row in df.iterrows():
        rows.append({
            'Test Reference':                  row['check_id'],
            'Description':                     row['description'],
            'Dimension':                       row['dimension'],
            'Severity':                        row['severity'],
            'Failing Records':                 int(row['failing']),
            'Total Assessed':                  int(row['total']),
            'Error Rate':                      f"{row['error_rate']:.1f}%",
            'RAG':                             row['rag'],
            'Comments':                        '',
            'Source System Cleansing Complete': '',
        })

    out_df = pd.DataFrame(rows)

    os.makedirs('trackers', exist_ok=True)
    out_path = os.path.join('trackers', f"{tab}_tracker_{house}.xlsx")

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='DQ Tracker')
        ws = writer.sheets['DQ Tracker']

        # Column widths
        col_widths = {
            'A': 30,  # Test Reference
            'B': 55,  # Description
            'C': 20,  # Dimension
            'D': 12,  # Severity
            'E': 18,  # Failing Records
            'F': 16,  # Total Assessed
            'G': 12,  # Error Rate
            'H': 8,   # RAG
            'I': 40,  # Comments
            'J': 35,  # Source System Cleansing Complete
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Header row styling
        for cell in ws[1]:
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[1].height = 30

        # Data rows
        sev_col  = out_df.columns.get_loc('Severity') + 1
        rag_col  = out_df.columns.get_loc('RAG') + 1
        fail_col = out_df.columns.get_loc('Failing Records') + 1

        for i, (_, row) in enumerate(out_df.iterrows(), start=2):
            ws.row_dimensions[i].height = 18
            for cell in ws[i]:
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=False)
                cell.font      = Font(size=10)

            # Severity colour
            sev_cell = ws.cell(row=i, column=sev_col)
            sev_fill = SEV_FILL.get(row['Severity'])
            if sev_fill:
                sev_cell.fill = PatternFill('solid', fgColor=sev_fill)
                sev_cell.font = Font(bold=True, size=10)

            # RAG colour
            rag_cell = ws.cell(row=i, column=rag_col)
            rag_fill = RAG_FILL.get(row['RAG'])
            if rag_fill:
                rag_cell.fill = PatternFill('solid', fgColor=rag_fill)
                rag_cell.font = Font(bold=True, size=10)

            # Bold failing count
            ws.cell(row=i, column=fail_col).font = Font(bold=True, size=10)

        # Freeze header row
        ws.freeze_panes = 'A2'

    print(f"Tracker saved: {out_path}  ({len(out_df)} checks with failures)")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate DQ tracker Excel file')
    parser.add_argument('tab',   help='Tab name: suppliers, customers, gl, assets')
    parser.add_argument('house', help='House: HOC or HOL')
    args = parser.parse_args()
    make_tracker(args.tab, args.house)
