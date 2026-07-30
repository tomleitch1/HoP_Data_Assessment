"""
Parliament Finance Systems Programme
Atamis DQ Export Generator
==========================
Generates ONE combined DQ tracker and one evidence file per check for the
Atamis domain, covering HOC, HOL, and Unknown together in the same files —
unlike every other domain's export tooling (generate_tracker.py /
generate_full_export.py), which produces a separate tracker/evidence set per
house. Atamis's own dashboard tab already treats 'Unknown' as a distinct,
non-house bucket rather than a third house (see CLAUDE.md), and per direct
request the export should mirror that: one tracker, one evidence file per
check, with a House column making clear which bucket each row belongs to,
saved into their own trackers/atamis/ folder rather than split by house.

Usage:
    python scripts/generate_atamis_export.py

Output:
    trackers/atamis/atamis_tracker.xlsx
    trackers/atamis/evidence/<Dimension>_<CHECK_ID>.xlsx
        (one per check with at least one failing record anywhere, combining
        failing records from every house that check has any in)
"""

import sys
import os

# Allow imports from project root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dashboard.data_engine import load_data, run_dq_analysis, get_failing_records
from dashboard.core.config import SCOPE_CONFIG

SCOPE_IDS = SCOPE_CONFIG['atamis']['scope_ids']

HEADER_FILL  = PatternFill('solid', fgColor='1E1528')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=10)
THIN_BORDER  = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
)
SEV_FILL = {
    'Critical': 'FFD7D7',
    'High':     'FFE8CC',
    'Medium':   'FFF3CC',
    'Low':      'E8F5E9',
}
RAG_FILL = {
    'Red':   'FFCCCC',
    'Amber': 'FFE5B4',
    'Green': 'CCFFCC',
}
# Distinct from SEV_FILL/RAG_FILL's palette so the eye doesn't confuse House
# with Severity/RAG at a glance.
HOUSE_FILL = {
    'HOC':     'D9EAD3',
    'HOL':     'F4CCCC',
    'Unknown': 'D9D9D9',
}


# ---------------------------------------------------------------------------
# Tracker
# ---------------------------------------------------------------------------

def _write_tracker(out_df: pd.DataFrame, out_path: str) -> None:
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='DQ Tracker')
        ws = writer.sheets['DQ Tracker']

        col_widths = {
            'A': 32, 'B': 10, 'C': 55, 'D': 20, 'E': 12,
            'F': 18, 'G': 16, 'H': 12, 'I': 8, 'J': 40, 'K': 35,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30

        house_col = out_df.columns.get_loc('House') + 1
        sev_col   = out_df.columns.get_loc('Severity') + 1
        rag_col   = out_df.columns.get_loc('RAG') + 1
        fail_col  = out_df.columns.get_loc('Failing Records') + 1

        for i, (_, row) in enumerate(out_df.iterrows(), start=2):
            ws.row_dimensions[i].height = 18
            for cell in ws[i]:
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=False)
                cell.font      = Font(size=10)

            house_fill = HOUSE_FILL.get(row['House'])
            if house_fill:
                hc = ws.cell(row=i, column=house_col)
                hc.fill = PatternFill('solid', fgColor=house_fill)
                hc.font = Font(bold=True, size=10)

            sev_fill = SEV_FILL.get(row['Severity'])
            if sev_fill:
                sc = ws.cell(row=i, column=sev_col)
                sc.fill = PatternFill('solid', fgColor=sev_fill)
                sc.font = Font(bold=True, size=10)

            rag_fill = RAG_FILL.get(row['RAG'])
            if rag_fill:
                rc = ws.cell(row=i, column=rag_col)
                rc.fill = PatternFill('solid', fgColor=rag_fill)
                rc.font = Font(bold=True, size=10)

            ws.cell(row=i, column=fail_col).font = Font(bold=True, size=10)

        ws.freeze_panes = 'A2'


def make_tracker(dq_results: pd.DataFrame, out_dir: str) -> None:
    df = dq_results[
        (dq_results['scope_id'].isin(SCOPE_IDS)) & (dq_results['failing'] > 0)
    ].copy()

    if df.empty:
        print("  No failing checks found — skipping tracker.")
        return

    # House ordered HOC/HOL first, Unknown last — a fixed reconciliation
    # bucket sorted at the end rather than interleaved.
    house_order = {'HOC': 0, 'HOL': 1, 'Unknown': 2}
    df['_house_sort'] = df['house'].map(house_order).fillna(9)
    df = df.sort_values(['_house_sort', 'scope_id', 'severity', 'check_id'])

    rows = []
    for _, row in df.iterrows():
        rows.append({
            'Test Reference':                  row['check_id'],
            'House':                           row['house'],
            'Description':                     row['description'],
            'Dimension':                       row['dimension'],
            'Severity':                        row['severity'],
            'Failing Records':                 int(row['failing']),
            'Total Assessed':                  int(row['total']),
            'Error Rate':                       f"{row['error_rate']:.1f}%",
            'RAG':                             row['rag'],
            'Comments':                        '',
            'Source System Cleansing Complete': '',
        })

    out_df   = pd.DataFrame(rows)
    out_path = os.path.join(out_dir, 'atamis_tracker.xlsx')
    _write_tracker(out_df, out_path)
    print(f"  Tracker saved: {out_path}  ({len(out_df)} check/house rows with failures)")


# ---------------------------------------------------------------------------
# Evidence files — one per check, combining every house that check applies to
# ---------------------------------------------------------------------------

def _style_evidence_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        sample = df[col_name].head(200) if not df.empty else df[col_name]
        max_len = max(len(str(col_name)), sample.astype(str).str.len().max() if not sample.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28

    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 16
        for cell in ws[i]:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.font      = Font(size=9)

    ws.freeze_panes = 'A2'


def _safe_filename(dimension: str, check_id: str) -> str:
    return f"{dimension}_{check_id}.xlsx"


def export_evidence(dq_results: pd.DataFrame, frames: dict, out_dir: str) -> None:
    failing_checks = dq_results[
        (dq_results['scope_id'].isin(SCOPE_IDS)) & (dq_results['failing'] > 0)
    ].copy()

    if failing_checks.empty:
        print("  No failing checks — no evidence files to write.")
        return

    written = 0
    for check_id, group in failing_checks.groupby('check_id'):
        dimension = group.iloc[0]['dimension']
        houses = group['house'].tolist()

        combined = []
        for house in houses:
            try:
                records = get_failing_records(check_id, house, frames, for_export=True)
            except Exception as exc:
                print(f"  [WARN] Could not retrieve records for {check_id} / {house}: {exc}")
                continue
            if records is None or records.empty:
                continue
            records = records.copy()
            # Always a clean single 'House' column with the value this
            # specific call was made for — some early-return blocks already
            # include their own lowercase 'house' column, others don't, so
            # relying on whatever's already there would be inconsistent.
            if 'house' in records.columns:
                records = records.drop(columns=['house'])
            records.insert(0, 'House', house)
            combined.append(records)

        if not combined:
            continue

        records_df = pd.concat(combined, ignore_index=True, sort=False)

        fname    = _safe_filename(dimension, check_id)
        out_path = os.path.join(out_dir, fname)

        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            meta_rows = [
                ['Check ID',                    check_id],
                ['Description',                 group.iloc[0]['description']],
                ['Dimension',                   dimension],
                ['Severity',                    group.iloc[0]['severity']],
                ['Houses included',             ', '.join(houses)],
                ['Total Failing (all houses)',  int(group['failing'].sum())],
                ['Total Assessed (all houses)', int(group['total'].sum())],
            ]
            meta_df = pd.DataFrame(meta_rows, columns=['Field', 'Value'])
            meta_df.to_excel(writer, index=False, sheet_name='Check Info')
            ws_meta = writer.sheets['Check Info']
            ws_meta.column_dimensions['A'].width = 26
            ws_meta.column_dimensions['B'].width = 70
            for cell in ws_meta[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            for i in range(2, ws_meta.max_row + 1):
                for cell in ws_meta[i]:
                    cell.border = THIN_BORDER
                    cell.font   = Font(size=10)
                ws_meta.cell(row=i, column=1).font = Font(bold=True, size=10)

            records_df.to_excel(writer, index=False, sheet_name='Failing Records')
            _style_evidence_sheet(writer.sheets['Failing Records'], records_df)

        written += 1

    print(f"  Evidence files saved: {written} files in {out_dir}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run() -> None:
    print("Loading data...")
    frames = load_data(tab='atamis')
    print("Running DQ analysis...")
    dq_results = run_dq_analysis(frames, tab='atamis')

    out_dir      = os.path.join('trackers', 'atamis')
    evidence_dir = os.path.join(out_dir, 'evidence')
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(evidence_dir, exist_ok=True)

    print("\n[1/2] Generating tracker...")
    make_tracker(dq_results, out_dir)

    print("\n[2/2] Generating evidence files...")
    export_evidence(dq_results, frames, evidence_dir)

    print("\nDone.")


if __name__ == '__main__':
    run()
